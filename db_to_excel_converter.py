#db_to_excel_converter.py
#Export database data to excel there is filtering option to specify which data should be imported and which one not
import os
import pandas as pd
from sqlalchemy import create_engine
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv("secret_keys.env")

# Define your connection parameters
username = "postgres"
password = os.getenv("DB_PSWD")
database = "mydb"
host = "localhost"
port = "5432"

# Create the SQLAlchemy engine for PostgreSQL using psycopg2 as the driver
engine = create_engine(f"postgresql+psycopg2://{username}:{password}@{host}:{port}/{database}")

# Define SQL queries to fetch data from the tables
query_invoices = "SELECT * FROM invoices;"
query_items = "SELECT * FROM invoice_items;"

# Define the Excel file path
excel_file_path = "/home/pavol/Documents/Python_Codes/OCR_Invoices/invoices.xlsx"

# Boolean flag: if True, import all rows; if False, apply filtering
import_all = True

# Specify filtering criteria (used only if import_all is False)
manual_ids = [None]  # Only include invoices with these IDs
range_start = None        # Also include invoices with IDs starting from 10
range_end = None          # up to and including 20

try:
    # Fetch the latest data from the database using the SQLAlchemy engine
    df_invoices_new = pd.read_sql(query_invoices, engine)
    df_items_new = pd.read_sql(query_items, engine)
    print("Data fetched successfully using SQLAlchemy engine.")

    # If not importing all rows, apply the filtering condition
    if not import_all:
        df_invoices_new = df_invoices_new[
            df_invoices_new['id'].isin(manual_ids) | 
            ((df_invoices_new['id'] >= range_start) & (df_invoices_new['id'] <= range_end))
        ]
        # Filter invoice items to include only those that belong to the filtered invoices.
        invoice_ids = df_invoices_new['id'].unique()
        df_items_new = df_items_new[df_items_new['invoice_id'].isin(invoice_ids)]
    else:
        print("Importing all rows without applying filters.")

    # Check if the Excel file already exists
    if os.path.exists(excel_file_path):
        print("Excel file exists. Checking for new rows...")

        # Read the existing sheets; if a sheet doesn't exist, use an empty DataFrame.
        try:
            df_invoices_existing = pd.read_excel(excel_file_path, sheet_name='Invoices')
        except Exception as e:
            print("Error reading 'Invoices' sheet, creating new:", e)
            df_invoices_existing = pd.DataFrame()

        try:
            df_items_existing = pd.read_excel(excel_file_path, sheet_name='Invoice Items')
        except Exception as e:
            print("Error reading 'Invoice Items' sheet, creating new:", e)
            df_items_existing = pd.DataFrame()

        # Identify new rows based on the unique identifier "id" (for invoices)
        if not df_invoices_existing.empty:
            new_invoices = df_invoices_new[~df_invoices_new['id'].isin(df_invoices_existing['id'])]
        else:
            new_invoices = df_invoices_new

        # For invoice items, use a similar approach based on the linking column "invoice_id"
        if not df_items_existing.empty:
            new_items = df_items_new[~df_items_new['invoice_id'].isin(df_items_existing['invoice_id'])]
        else:
            new_items = df_items_new

        # Append new rows if there are any
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Append new invoices
            if not new_invoices.empty:
                startrow = writer.sheets['Invoices'].max_row
                new_invoices.to_excel(writer, sheet_name='Invoices', startrow=startrow, index=False, header=False)
                print(f"Appended {len(new_invoices)} new invoice rows.")
            else:
                print("No new invoice rows to append.")

            # Append new invoice items
            if not new_items.empty:
                startrow = writer.sheets['Invoice Items'].max_row
                new_items.to_excel(writer, sheet_name='Invoice Items', startrow=startrow, index=False, header=False)
                print(f"Appended {len(new_items)} new invoice item rows.")
            else:
                print("No new invoice item rows to append.")
    else:
        # If the file does not exist, create a new file with both sheets
        with pd.ExcelWriter(excel_file_path) as writer:
            df_invoices_new.to_excel(writer, sheet_name='Invoices', index=False)
            df_items_new.to_excel(writer, sheet_name='Invoice Items', index=False)
        print("Excel file created with initial data.")

except Exception as e:
    print("An error occurred:", e)
